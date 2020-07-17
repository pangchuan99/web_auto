from openpyxl import load_workbook


wb=load_workbook("小黄.xlsx")
sheet=wb['python自动化学习']

res=sheet.cell(1,1).value
print(res)


